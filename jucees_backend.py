##########################################################################
# Esse Script só funciona dentro da rede interna do órgão conveniado
##########################################################################


from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchWindowException
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import pandas as pd
import os
from os import listdir
#from webdriver_manager.chrome import ChromeDriverManager




class jucees:
    def __init__(self) -> None:
        pass

    def expand_shadow_element(self, driver, element):
        shadow_root = driver.execute_script('return arguments[0].shadowRoot', element)
        return shadow_root

    def obtem_nome_do_download(self, driver):
        driver.implicitly_wait(60)  # in seconds
        driver.get('chrome://downloads')
        sleep(2)
        root1 = driver.find_element(By.TAG_NAME, 'downloads-manager')
        shadow_root1 = self.expand_shadow_element(driver, root1)
        sleep(2)
        root2 = shadow_root1.find_element(By.ID, 'frb0')
        shadow_root2 = self.expand_shadow_element(driver, root2)
        arquivo_baixado = ''
        tentativas = 80
        while (arquivo_baixado == ''):
            arquivo_baixado = shadow_root2.find_element(By.TAG_NAME, 'a').text  
            tentativas = tentativas - 1
            sleep(1)
            if tentativas == 0:
                break
        return arquivo_baixado

    def cnpj_sem_mascara(self, cnpj):
        cnpj = cnpj.replace("/", "")
        cnpj = cnpj.replace("-", "")
        cnpj = cnpj.replace(".", "")
        cnpj = cnpj.replace(" ", "")
        return cnpj
    
    def get_html(self, elemento):
        html = elemento.get_attribute('innerHTML')
        soup = BeautifulSoup(html, 'lxml')
        return soup
    
    def carregou_pagina(self, driver_navegador, by_content, by=By.ID, timeout = 8, to_sleep = 0, loop = True):
        fica_no_loop = True
        while (fica_no_loop):
            fica_no_loop = loop
            try:
                wait = WebDriverWait(driver_navegador, timeout, poll_frequency=0.5)
                element = wait.until(EC.presence_of_element_located((by, by_content)))
                fica_no_loop = False
            except TimeoutException:
                print("A página ainda não carregou")
                if not fica_no_loop: 
                    return False
            except NoSuchWindowException:
                sleep(1)
            if (to_sleep > 0):
                sleep(to_sleep)
            if not fica_no_loop:
                return True

    def obtem_nome_do_download2(self, dir):
        x = listdir(dir)

        while len(x) == 0:
            sleep(1)
            x = listdir(dir)
        
        for linha in x:
            while linha.rpartition('.')[2] != 'pdf':
               sleep(1)
               x = listdir(dir)
               linha=x[0]
            return linha


    def scrap(self, lista):
        print("ATENÇÃO: Esse Script só funciona dentro da rede interna do órgão conveniado ou usando VPN \ workspace deste órgão. \n\n")
        alvos = lista
        url = 'https://www.jucees.es.gov.br/certidaoweb/index.php' # página principal
        url2 = 'https://www.jucees.es.gov.br/certidaoweb/index.php?acao_usuario=empresas&cont_sel=177&guia_sel=2&listar=1' # página de consulta
        dir_trabalho = os.path.dirname(os.path.realpath(__file__))
        dir_download = dir_trabalho + '\jucees_download'
        dir_resultado = dir_trabalho + '\jucees_resultado'


        if not os.path.exists(dir_download):
            print("Criando o diretório: " + dir_download)
            os.makedirs(dir_download)

        if not os.path.exists(dir_resultado):
            print("Criando o diretório: " + dir_resultado)
            os.makedirs(dir_resultado)
            
        chrome_options = Options()
        chrome_options.add_experimental_option('prefs', {"plugins.always_open_pdf_externally": True, "download.default_directory" : dir_download  }) #Não mostra o PDF no navegador, só o link para baixar. 
        # driver = webdriver.Chrome("chromedriver.exe", options=chromeOptions)
        # driver = webdriver.Chrome(ChromeDriverManager(path = dirname).install(), options=chromeOptions) Esta linha está desabilitada devido ao problema de atualização.
        # O código novo que contornou o problema segue abaixo:
        
        ###########################################################################################################
        #IMPORTANTE: O selenium deve ser 4.11.2 ou superior.
        from selenium.webdriver.chrome.service import Service
        service = Service()

        driver = webdriver.Chrome(service=service, options=chrome_options) 
        ###########################################################################################################
        # fim do código novo
       
        wait = WebDriverWait(driver, 60)
        

        # Imprime a hora de início
        now = datetime.now()
        print("Hora de ínicio: " + str(now.hour) + ":" + str(now.minute) + ":" + str(now.second)) 

        # Acessa o site principal do JUCEES, na URL do acesso restrito
        driver.get(url)
        driver.implicitly_wait(5)  # in seconds
        self.carregou_pagina(driver, "menusup", By.ID, timeout = 5, loop = True)
        driver.get(url2)
        self.carregou_pagina(driver, 'nrCNPJ1', By.NAME, timeout = 5, loop = True)
        sleep(2)


        campo_cnpj = driver.find_element(By.NAME, 'nrCNPJ1')  
        janela_consulta = driver.current_window_handle

        for alvo in alvos:
            alvo_num = self.cnpj_sem_mascara(alvo)
            alvo = alvo_num
            print("Buscando o CNPJ: " + alvo)
            campo_cnpj.clear()
            campo_cnpj.send_keys(alvo)
            sleep(1)
            botao_lupa = driver.find_element(By.XPATH, '//*[@id="corpo"]/div/form[1]/table/tbody/tr[3]/td[2]/table/tbody/tr/td[2]/a')
            botao_lupa.click() # tratar caso de nao encontrar o CNPJ
            sleep(5)   # Mudar para checar por um tempo.

            nome_empresa = driver.find_element(By.XPATH, '/html/body/div/div[3]/div/form[1]/table/tbody/tr[4]/td[2]/table/tbody/tr/td')
            achou_empresa = nome_empresa.get_attribute('innerHTML').find('readonly')
            if achou_empresa < 0:
                pass

            else:
                botao_exibir_dados = driver.find_element(By.XPATH, '//*[@id="corpo"]/div/form[1]/table/tbody/tr[5]/td/table/tbody/tr/td/input[1]')
                botao_exibir_dados.click()
                sleep(5)

                # Acessa os dados da empresa, tira o screenshot, obtem nome da empresa
                wait.until(EC.number_of_windows_to_be(2))
                for window_handle in driver.window_handles:
                    if window_handle != janela_consulta:
                        driver.switch_to.window(window_handle)
                        janela_download = driver.current_window_handle
                        break
                frame_links = driver.find_element(By.NAME, 'vinculos')
                driver.switch_to.frame(frame_links)
                
                nome_empresa = driver.find_element(By.XPATH, '//*[@id="corpo"]/div/table[1]/tbody/tr/td[1]/table/tbody/tr[2]/td').text
                nome_empresa_reduzido = nome_empresa.replace(" ", "")[:8]
                cnpj = driver.find_element(By.XPATH, '//*[@id="corpo"]/div/table[1]/tbody/tr/td[2]/table/tbody/tr[2]/td').text
                nucleo_filename = dir_resultado + '\\' + alvo_num + ' - ' + nome_empresa_reduzido 
                driver.save_screenshot(str(nucleo_filename + '-dados.png'))

                # Prepara o Excel e coleta os dados da empresa
                wb = Workbook() # Cria a pasta de trabalho
                ws = wb.active # Pega a planilha ativa
                ws.title = "Dados da empresa"
                linha = ['Nome da empresa', 'CNPJ']
                ws.append(linha)
                linha = [nome_empresa, cnpj]
                ws.append(linha)

                # Pega a lista de Atividades
                link_atividades = driver.find_element(By.XPATH, '/html/body/div/div[2]/div/div/table/tbody/tr/td[2]/b/a')
                link_atividades.click()
                sleep(2)
                driver.save_screenshot(str(dir_resultado + '//' + alvo_num + '-atividades.png'))
                tab_lista_ativ = self.get_html(driver.find_element(By.XPATH, '//*[@id="corpo"]/div/table[2]/tbody/tr/td/table/tbody/tr[2]/td')).find_all('li')
                ws_atividades = wb.create_sheet('Atividades')
                ws_atividades.append(['Atividades'])
                for ativ in tab_lista_ativ:
                    ws_atividades.append([ativ.text])

                # Pega a lista de Pessoas
                ws_pessoas = wb.create_sheet('Pessoas')
                ws_pessoas.append(['CPF / CNPJ', 'Endereço', 'Participação', 'Em %', 'Vínculo', 'Entrada', 'Saida'])
                driver.find_element(By.XPATH, '/html/body/div/div[2]/div/div/table/tbody/tr/td[3]/b/a').click()
                sleep(2)
                elem_tabela_pessoa = driver.find_element(By.XPATH, '//*[@id="corpo"]/div/table[2]/tbody')
                html = self.get_html(elem_tabela_pessoa)
                dados_tab_pessoa = html.find_all('td')
                num_celulas = len(dados_tab_pessoa)
                celula = 8
                while (celula < num_celulas):
                    linha = []
                    for aux in range (7):
                        if (celula < num_celulas):
                            linha.append(dados_tab_pessoa[celula].text)
                        celula += 1
                    ws_pessoas.append(linha)
                
                # Formata a largura das células (lista de pessoas) 
                num_linhas = int((num_celulas-1) / 7)
                ws_pessoas.column_dimensions['A'].width = 18
                ws_pessoas.column_dimensions['B'].width = 180
                ws_pessoas.column_dimensions['C'].width = 15
                ws_pessoas.column_dimensions['E'].width = 20
                ws_pessoas.column_dimensions['F'].width = 12
                ws_pessoas.column_dimensions['G'].width = 10

                # Coloca as celulas da coluna B em "quebrar linhas automat."
                for linha in range(1, num_linhas+1, 1):
                    endereco_celula = str('B' + str(linha))
                    cell = ws_pessoas[endereco_celula]
                    alignment = Alignment(wrapText=True)
                    cell.alignment = alignment



                # Pega a lista de filiais - testar se não tiver filiais. 
                ws_filiais = wb.create_sheet('Filiais')
                ws_filiais.append(['NIRE', 'CNPJ', 'Endereço', 'Situação'])
                
                driver.find_element(By.XPATH, '/html/body/div/div[2]/div/div/table/tbody/tr/td[4]/b/a').click()
                sleep(2)
                elem_tabela_filiais = driver.find_element(By.XPATH, '//*[@id="corpo"]/div/table[2]/tbody')
                html_filiais = self.get_html(elem_tabela_filiais)
                dados_tab_filiais = html_filiais.find_all('td')
                num_celulas = len(dados_tab_filiais)
                celula = 5
                while (celula < num_celulas):
                    linha = []
                    for aux in range (4):
                        if (celula == 5) and (dados_tab_filiais[5].text == 'INEXISTENTE'):
                            break
                        linha.append(dados_tab_filiais[celula].text)
                        celula += 1
                    if celula != 5: 
                        ws_filiais.append(linha)
                    else: 
                        break
                
                # Formata a largura das células (lista de pessoas) e coloca as celulas da coluna B em "quebrar linhas automat."
                num_linhas = int((num_celulas-1) / 4)
                ws_filiais.column_dimensions['A'].width = 15
                ws_filiais.column_dimensions['B'].width = 20
                ws_filiais.column_dimensions['C'].width = 180
                ws_filiais.column_dimensions['D'].width = 20

                # Coloca as celulas da coluna C em "quebrar linhas automat."
                for linha in range(1, num_linhas+1, 1):
                    endereco_celula = str('C' + str(linha))
                    cell = ws_pessoas[endereco_celula]
                    alignment = Alignment(wrapText=True)
                    cell.alignment = alignment

                link_historico = driver.find_element(By.XPATH, '/html/body/div/div[2]/div/div/table/tbody/tr/td[5]/b/a')
                link_historico.click()
                sleep(2)
                elem_tabela_historico = driver.find_element(By.XPATH, '//*[@id="corpo"]/div/table[2]')
                html = elem_tabela_historico.get_attribute('outerHTML')
                
                df = pd.read_html(html)
                df2 = df[0][8:].drop(columns = [5])
                df2.columns = ['Data', 'Ato', 'Evento', 'Arquivamento', 'Protocolo']
                ws_historico = wb.create_sheet('Historico')
                for linhas_historico in dataframe_to_rows(df2, index=False, header=True):
                    ws_historico.append(linhas_historico)
                
                ws_historico.column_dimensions['A'].width = 15
                ws_historico.column_dimensions['C'].width = 100
                ws_historico.column_dimensions['D'].width = 15
                ws_historico.column_dimensions['E'].width = 15

                try:
                    wb.save(nucleo_filename + '.xlsx')
                    print("Arquivo salvo: " + nucleo_filename + ".xlsx")
                except:
                    try:
                        wb.save(nucleo_filename + '2.xlsx')
                    except:
                        print("Não foi possível salvar o arquivo " + nucleo_filename + ".xlsx")
                                
                wb.close

                # Faz o download dos PDFs
                num_linhas_historico = df2.shape[0]
                for linha in range (0, num_linhas_historico, 1):
                    # Usa a URL para acesso direto, sem clicar no link. 
                    if pd.isna(df2['Protocolo'].values[linha]):
                        pass
                    else:
                        url_pdf = 'https://apps.jucees.es.gov.br/certidaoweb/imagem.php?nire=&protocolo=' + df2['Protocolo'].values[linha] + '&ato=' + df2['Ato'].values[linha] + '&read=1'
                        driver.get(url_pdf)

                        sleep(3)
                        protocolo = df2['Protocolo'].values[linha]
                        print("Download do " + str(linha+1) + "º arquivo do protocolo " + str(protocolo) + " do CNPJ " + str(alvo))
                        iframe_pdf = driver.find_element(By.ID, 'pdf')
                        driver.switch_to.frame(iframe_pdf)
                        
                        driver.find_element(By.XPATH, '//*[@id="open-button"]').click()
                        filename_novo = nucleo_filename + "_" + protocolo + ".pdf"
                        filename_antigo_sem_path = self.obtem_nome_do_download2(dir_download)
                        filename_antigo = dir_download + '\\' + filename_antigo_sem_path

                        while (not os.path.exists(filename_antigo)):
                            sleep(1)
                        

                        try: 
                            os.rename(filename_antigo, filename_novo)
                        except: 
                            #Se o arquivo existir, baixa de novo acrescentando uma letra ao nome. Tenta até 4 letras. 
                            filename_novo_a = dir_resultado + "\\" + alvo_num + "_" + protocolo + "a.pdf"
                            filename_novo_b = dir_resultado + "\\" + alvo_num + "_" + protocolo + "b.pdf"
                            filename_novo_c = dir_resultado + "\\" + alvo_num + "_" + protocolo + "c.pdf"
                            filename_novo_d = dir_resultado + "\\" + alvo_num + "_" + protocolo + "d.pdf"
                            if not os.path.exists(filename_novo_a):
                                os.rename(filename_antigo, filename_novo_a)
                            elif not os.path.exists(filename_novo_b):
                                os.rename(filename_antigo, filename_novo_b)
                            elif not os.path.exists(filename_novo_c):
                                os.rename(filename_antigo, filename_novo_c)
                            elif not os.path.exists(filename_novo_d):
                                os.rename(filename_antigo, filename_novo_d)
                            else:
                                pass
                        
                        # Espera que o arquivo seja completamente movido para a pasta nova
                        while len(listdir(dir_download)) > 0:
                            sleep(1)
        
                sleep(1)
                driver.switch_to.window(janela_download)
                driver.execute_script('window.close()')
                driver.switch_to.window(janela_consulta)

        fim = datetime.now()
        print("Hora de fim: " + str(fim.hour) + ":" + str(fim.minute) + ":" + str(fim.second)) 
        print ("Concluído")
        driver.close()
    

